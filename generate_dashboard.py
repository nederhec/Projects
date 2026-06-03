"""
Generate Dashboard_Cargos_GBS.xlsx from HR data.
"""
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
warnings.filterwarnings("ignore")

# ── colours ──────────────────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"
LIGHT_BLUE  = "BDD7EE"
MID_BLUE    = "2E75B6"
ALT_ROW     = "DEEAF1"
WHITE       = "FFFFFF"
ORANGE      = "F4B942"
GREEN       = "70AD47"
RED_LIGHT   = "FFE0E0"

# ── helper styles ─────────────────────────────────────────────────────────────
def hdr_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_header_row(ws, row_idx, num_cols, bg=DARK_BLUE, fg=WHITE, bold=True):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = hdr_fill(bg)
        cell.font = Font(color=fg, bold=bold, name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()

def apply_data_row(ws, row_idx, num_cols, alternate=False):
    bg = ALT_ROW if alternate else WHITE
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = hdr_fill(bg)
        cell.font = Font(name="Calibri", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

# ── load data ─────────────────────────────────────────────────────────────────
SOURCE = "/root/.claude/uploads/39518091-0fb9-4665-a41a-ed8e752c52fd/88b7cf84-Cargos_GBS.xlsx"
df = pd.read_excel(SOURCE)

wb = Workbook()
wb.remove(wb.active)   # remove default sheet

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 – KPIs Executivos
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("KPIs Executivos")
ws1.sheet_view.showGridLines = False

# Title
ws1.merge_cells("A1:F1")
title = ws1["A1"]
title.value = "KPIs EXECUTIVOS – CARGOS GBS"
title.fill = hdr_fill(DARK_BLUE)
title.font = Font(color=WHITE, bold=True, size=16, name="Calibri")
title.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 36

# Sub-header
ws1.merge_cells("A2:F2")
sub = ws1["A2"]
sub.value = "Estrutura de Cargos – Visão Geral"
sub.fill = hdr_fill(MID_BLUE)
sub.font = Font(color=WHITE, bold=False, size=11, name="Calibri")
sub.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[2].height = 22

ws1.append([])  # blank row 3

# Helper to write a KPI box (label row + value row)
def write_kpi(ws, start_row, start_col, label, value, note="", label_color=MID_BLUE):
    # label cell (merged 2 cols)
    ws.merge_cells(
        start_row=start_row, start_column=start_col,
        end_row=start_row,   end_column=start_col + 1
    )
    lbl = ws.cell(row=start_row, column=start_col)
    lbl.value = label
    lbl.fill = hdr_fill(label_color)
    lbl.font = Font(color=WHITE, bold=True, size=10, name="Calibri")
    lbl.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lbl.border = thin_border()
    ws.row_dimensions[start_row].height = 30

    # value cell
    ws.merge_cells(
        start_row=start_row + 1, start_column=start_col,
        end_row=start_row + 1,   end_column=start_col + 1
    )
    val = ws.cell(row=start_row + 1, column=start_col)
    val.value = value
    val.fill = hdr_fill(LIGHT_BLUE)
    val.font = Font(bold=True, size=18, name="Calibri", color=DARK_BLUE)
    val.alignment = Alignment(horizontal="center", vertical="center")
    val.border = thin_border()
    ws.row_dimensions[start_row + 1].height = 40

    if note:
        ws.merge_cells(
            start_row=start_row + 2, start_column=start_col,
            end_row=start_row + 2,   end_column=start_col + 1
        )
        n = ws.cell(row=start_row + 2, column=start_col)
        n.value = note
        n.font = Font(size=9, italic=True, name="Calibri", color="555555")
        n.alignment = Alignment(horizontal="center")
        ws.row_dimensions[start_row + 2].height = 16

total       = len(df)
ativos      = (df["SITUAÇÃO"] == "ATIVO").sum()
inativos    = (df["SITUAÇÃO"] == "INATIVO").sum()
empresas    = df["EMPRESA"].nunique()
familias    = df["FAMÍLIA GBS"].nunique()
hierarquias = df["GRUPO HIERARQUICO"].nunique()
sem_mercer  = df["Cod. Mercer 2026"].isna().sum()
corp        = (df["GRUPO"] == "Corporativo").sum()
neg_op      = (df["GRUPO"] == "Negócios & operações").sum()

kpis = [
    ("Total de Cargos",             total,        "",                                       4, MID_BLUE),
    ("Cargos Ativos",               ativos,       f"{ativos/total:.1%} do total",           4, GREEN),
    ("Cargos Inativos",             inativos,     f"{inativos/total:.1%} do total",         4, "C00000"),
    ("Total de Empresas",           empresas,     "",                                       8, MID_BLUE),
    ("Famílias GBS",                familias,     "",                                       8, MID_BLUE),
    ("Grupos Hierárquicos",         hierarquias,  "",                                       8, MID_BLUE),
    ("Sem Cod. Mercer 2026",        sem_mercer,   f"{sem_mercer/total:.1%} dos cargos",    12, ORANGE),
    ("Cargos Corporativos",         corp,         f"{corp/total:.1%} do total",            12, MID_BLUE),
    ("Cargos Negócios & Operações", neg_op,       f"{neg_op/total:.1%} do total",          12, MID_BLUE),
]

# Lay out 3 KPIs per row, 2 cols each (cols A-B, D-E, G-H)
col_offsets = [1, 4, 7]
row_start = 4
for idx, (label, value, note, _, color) in enumerate(kpis):
    r = row_start + (idx // 3) * 4
    c = col_offsets[idx % 3]
    write_kpi(ws1, r, c, label, value, note, color)

# Distribution table at bottom
dist_row = row_start + 3 * 4 + 2
ws1.merge_cells(f"A{dist_row}:F{dist_row}")
h = ws1.cell(row=dist_row, column=1)
h.value = "Distribuição por GRUPO"
h.fill = hdr_fill(DARK_BLUE)
h.font = Font(color=WHITE, bold=True, size=12, name="Calibri")
h.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[dist_row].height = 28

dist_row += 1
headers = ["GRUPO", "Total", "Ativos", "Inativos", "% Ativos", "% do Total"]
for ci, h_text in enumerate(headers, 1):
    c = ws1.cell(row=dist_row, column=ci)
    c.value = h_text
    c.fill = hdr_fill(MID_BLUE)
    c.font = Font(color=WHITE, bold=True, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border()
ws1.row_dimensions[dist_row].height = 24

grupos = df.groupby("GRUPO")
for i, (grp, gdf) in enumerate(grupos):
    dist_row += 1
    tot_g  = len(gdf)
    ativ_g = (gdf["SITUAÇÃO"] == "ATIVO").sum()
    inat_g = (gdf["SITUAÇÃO"] == "INATIVO").sum()
    row_data = [grp, tot_g, ativ_g, inat_g, f"{ativ_g/tot_g:.1%}", f"{tot_g/total:.1%}"]
    for ci, val in enumerate(row_data, 1):
        c = ws1.cell(row=dist_row, column=ci)
        c.value = val
        c.fill = hdr_fill(ALT_ROW if i % 2 else WHITE)
        c.font = Font(name="Calibri", size=10)
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border()
    ws1.row_dimensions[dist_row].height = 20

# column widths
for col_letter, width in zip("ABCDEFG", [30, 6, 14, 14, 14, 14, 14]):
    ws1.column_dimensions[col_letter].width = width

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 – Distribuição por Grade
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Distribuição por Grade")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:E1")
t = ws2["A1"]
t.value = "DISTRIBUIÇÃO POR GRADE"
t.fill = hdr_fill(DARK_BLUE)
t.font = Font(color=WHITE, bold=True, size=14, name="Calibri")
t.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 32

headers2 = ["Grade", "Ativos", "Inativos", "Total", "% do Total"]
for ci, h in enumerate(headers2, 1):
    c = ws2.cell(row=2, column=ci)
    c.value = h
    c.fill = hdr_fill(MID_BLUE)
    c.font = Font(color=WHITE, bold=True, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border()
ws2.row_dimensions[2].height = 24

all_grades = sorted(df["GRADE"].dropna().unique())
grade_rows = []
for i, grade in enumerate(all_grades):
    gdf   = df[df["GRADE"] == grade]
    ativ  = (gdf["SITUAÇÃO"] == "ATIVO").sum()
    inat  = (gdf["SITUAÇÃO"] == "INATIVO").sum()
    tot   = ativ + inat
    pct   = tot / total
    grade_rows.append([int(grade), ativ, inat, tot, pct])
    row_idx = i + 3
    row_data = [int(grade), ativ, inat, tot, f"{pct:.1%}"]
    for ci, val in enumerate(row_data, 1):
        c = ws2.cell(row=row_idx, column=ci)
        c.value = val
        c.fill = hdr_fill(ALT_ROW if i % 2 else WHITE)
        c.font = Font(name="Calibri", size=10)
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border()
    ws2.row_dimensions[row_idx].height = 20

# totals row
tot_row = len(all_grades) + 3
ws2.cell(row=tot_row, column=1).value = "TOTAL"
ws2.cell(row=tot_row, column=2).value = ativos
ws2.cell(row=tot_row, column=3).value = inativos
ws2.cell(row=tot_row, column=4).value = total
ws2.cell(row=tot_row, column=5).value = "100,0%"
apply_header_row(ws2, tot_row, 5, DARK_BLUE)

# Bar chart – ativos por grade
chart2 = BarChart()
chart2.type     = "col"
chart2.title    = "Cargos Ativos por Grade"
chart2.y_axis.title = "Quantidade"
chart2.x_axis.title = "Grade"
chart2.style    = 10
chart2.width    = 18
chart2.height   = 12

data_ref   = Reference(ws2, min_col=2, max_col=2, min_row=2, max_row=len(all_grades) + 2)
cats_ref   = Reference(ws2, min_col=1, min_row=3, max_row=len(all_grades) + 2)
chart2.add_data(data_ref, titles_from_data=True)
chart2.set_categories(cats_ref)
chart2.series[0].graphicalProperties.solidFill = MID_BLUE
ws2.add_chart(chart2, "G2")

for col_letter, width in zip("ABCDE", [10, 12, 12, 12, 14]):
    ws2.column_dimensions[col_letter].width = width

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 – Análise por Família GBS
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Análise por Família GBS")
ws3.sheet_view.showGridLines = False

hier_order = ["GA", "GS", "Coordenador", "Supervisor", "Especialista", "Analista", "Assistente", "Auxiliar"]
hier_present = [h for h in hier_order if h in df["GRUPO HIERARQUICO"].unique()]

pivot = df.groupby(["FAMÍLIA GBS", "GRUPO HIERARQUICO"]).size().unstack(fill_value=0)
# reorder columns
cols_ordered = [h for h in hier_present if h in pivot.columns]
pivot = pivot[cols_ordered]
pivot["TOTAL"] = pivot.sum(axis=1)

ws3.merge_cells(f"A1:{get_column_letter(len(cols_ordered) + 3)}1")
t = ws3["A1"]
t.value = "ANÁLISE POR FAMÍLIA GBS × GRUPO HIERÁRQUICO"
t.fill = hdr_fill(DARK_BLUE)
t.font = Font(color=WHITE, bold=True, size=14, name="Calibri")
t.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 32

# header row
header_row = ["FAMÍLIA GBS"] + cols_ordered + ["TOTAL"]
for ci, h in enumerate(header_row, 1):
    c = ws3.cell(row=2, column=ci)
    c.value = h
    c.fill = hdr_fill(MID_BLUE)
    c.font = Font(color=WHITE, bold=True, name="Calibri", size=10)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin_border()
ws3.row_dimensions[2].height = 30

for i, (familia, row_data) in enumerate(pivot.iterrows()):
    r = i + 3
    ws3.cell(row=r, column=1).value = familia
    ws3.cell(row=r, column=1).font = Font(name="Calibri", size=10, bold=True)
    ws3.cell(row=r, column=1).fill = hdr_fill(ALT_ROW if i % 2 else WHITE)
    ws3.cell(row=r, column=1).border = thin_border()
    for ci, val in enumerate(row_data.values, 2):
        c = ws3.cell(row=r, column=ci)
        c.value = int(val)
        c.fill = hdr_fill(ALT_ROW if i % 2 else WHITE)
        c.font = Font(name="Calibri", size=10)
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border()
    ws3.row_dimensions[r].height = 20

# Grand total
gt_row = len(pivot) + 3
ws3.cell(row=gt_row, column=1).value = "GRAND TOTAL"
totals = [pivot[c].sum() for c in cols_ordered] + [pivot["TOTAL"].sum()]
for ci, val in enumerate(totals, 2):
    ws3.cell(row=gt_row, column=ci).value = int(val)
apply_header_row(ws3, gt_row, len(header_row), DARK_BLUE)

ws3.column_dimensions["A"].width = 32
for col_idx in range(2, len(header_row) + 2):
    ws3.column_dimensions[get_column_letter(col_idx)].width = 14

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 – Análise por Empresa
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Análise por Empresa")
ws4.sheet_view.showGridLines = False

ws4.merge_cells("A1:G1")
t = ws4["A1"]
t.value = "ANÁLISE POR EMPRESA"
t.fill = hdr_fill(DARK_BLUE)
t.font = Font(color=WHITE, bold=True, size=14, name="Calibri")
t.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 32

headers4 = ["EMPRESA", "Total Cargos", "Ativos", "Inativos", "% Ativo", "Grades Distintos", "Famílias GBS Distintas"]
for ci, h in enumerate(headers4, 1):
    c = ws4.cell(row=2, column=ci)
    c.value = h
    c.fill = hdr_fill(MID_BLUE)
    c.font = Font(color=WHITE, bold=True, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin_border()
ws4.row_dimensions[2].height = 28

empresas_sorted = df.groupby("EMPRESA").size().sort_values(ascending=False).index
for i, emp in enumerate(empresas_sorted):
    edf    = df[df["EMPRESA"] == emp]
    tot_e  = len(edf)
    ativ_e = (edf["SITUAÇÃO"] == "ATIVO").sum()
    inat_e = (edf["SITUAÇÃO"] == "INATIVO").sum()
    pct_a  = ativ_e / tot_e
    grades = edf["GRADE"].nunique()
    fams   = edf["FAMÍLIA GBS"].nunique()
    r = i + 3
    row_data = [emp, tot_e, ativ_e, inat_e, f"{pct_a:.1%}", grades, fams]
    for ci, val in enumerate(row_data, 1):
        c = ws4.cell(row=r, column=ci)
        c.value = val
        c.fill = hdr_fill(ALT_ROW if i % 2 else WHITE)
        c.font = Font(name="Calibri", size=10)
        c.alignment = Alignment(horizontal="center" if ci > 1 else "left", vertical="center")
        c.border = thin_border()
    ws4.row_dimensions[r].height = 20

# totals
last_r = len(list(empresas_sorted)) + 3
ws4.cell(row=last_r, column=1).value = "TOTAL"
ws4.cell(row=last_r, column=2).value = total
ws4.cell(row=last_r, column=3).value = ativos
ws4.cell(row=last_r, column=4).value = inativos
ws4.cell(row=last_r, column=5).value = f"{ativos/total:.1%}"
ws4.cell(row=last_r, column=6).value = df["GRADE"].nunique()
ws4.cell(row=last_r, column=7).value = df["FAMÍLIA GBS"].nunique()
apply_header_row(ws4, last_r, 7, DARK_BLUE)

ws4.column_dimensions["A"].width = 20
for col_letter, w in zip("BCDEFG", [14, 10, 10, 12, 18, 22]):
    ws4.column_dimensions[col_letter].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 – Benchmarking Mercer
# ══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Benchmarking Mercer")
ws5.sheet_view.showGridLines = False

ws5.merge_cells("A1:F1")
t = ws5["A1"]
t.value = "BENCHMARKING MERCER – COBERTURA DE CARGOS"
t.fill = hdr_fill(DARK_BLUE)
t.font = Font(color=WHITE, bold=True, size=14, name="Calibri")
t.alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[1].height = 32

com_2025    = df["Cod. Mercer 2025"].notna().sum()
com_2026    = df["Cod. Mercer 2026"].notna().sum()
sem_2026    = df["Cod. Mercer 2026"].isna().sum()
cobertura   = com_2026 / total

summary_data = [
    ("Cargos com Cod. Mercer 2025",  com_2025,  f"{com_2025/total:.1%}"),
    ("Cargos com Cod. Mercer 2026",  com_2026,  f"{com_2026/total:.1%}"),
    ("Cargos SEM Cod. Mercer 2026",  sem_2026,  f"{sem_2026/total:.1%}"),
    ("% Cobertura Mercer 2026",      f"{cobertura:.1%}", ""),
]

ws5.cell(row=2, column=1).value = "Indicador"
ws5.cell(row=2, column=2).value = "Quantidade"
ws5.cell(row=2, column=3).value = "% do Total"
apply_header_row(ws5, 2, 3, MID_BLUE)

for i, (label, qty, pct) in enumerate(summary_data):
    r = i + 3
    is_alt = i % 2 == 0
    for ci, val in enumerate([label, qty, pct], 1):
        c = ws5.cell(row=r, column=ci)
        c.value = val
        bg = RED_LIGHT if "SEM" in label else (ALT_ROW if is_alt else WHITE)
        c.fill = hdr_fill(bg)
        c.font = Font(name="Calibri", size=10,
                      bold=("SEM" in label or "Cobertura" in label))
        c.alignment = Alignment(horizontal="center" if ci > 1 else "left")
        c.border = thin_border()
    ws5.row_dimensions[r].height = 20

# Gap list
ws5.merge_cells("A8:F8")
gap_hdr = ws5["A8"]
gap_hdr.value = "CARGOS SEM COD. MERCER 2026 (lista detalhada)"
gap_hdr.fill = hdr_fill(ORANGE)
gap_hdr.font = Font(bold=True, name="Calibri", size=11, color=DARK_BLUE)
gap_hdr.alignment = Alignment(horizontal="center")
ws5.row_dimensions[8].height = 26

gap_cols = ["CARGO", "EMPRESA", "GRADE", "FAMÍLIA GBS", "GRUPO HIERARQUICO", "Cod. Mercer 2025"]
for ci, h in enumerate(gap_cols, 1):
    c = ws5.cell(row=9, column=ci)
    c.value = h
    c.fill = hdr_fill(MID_BLUE)
    c.font = Font(color=WHITE, bold=True, name="Calibri")
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    c.border = thin_border()
ws5.row_dimensions[9].height = 24

gap_df = df[df["Cod. Mercer 2026"].isna()][gap_cols].sort_values(["EMPRESA", "GRADE"])
for i, (_, row) in enumerate(gap_df.iterrows()):
    r = i + 10
    for ci, col in enumerate(gap_cols, 1):
        c = ws5.cell(row=r, column=ci)
        c.value = row[col]
        c.fill = hdr_fill(ALT_ROW if i % 2 else WHITE)
        c.font = Font(name="Calibri", size=10)
        c.alignment = Alignment(horizontal="center" if ci > 1 else "left", wrap_text=True)
        c.border = thin_border()
    ws5.row_dimensions[r].height = 20

ws5.column_dimensions["A"].width = 45
ws5.column_dimensions["B"].width = 18
ws5.column_dimensions["C"].width = 10
ws5.column_dimensions["D"].width = 30
ws5.column_dimensions["E"].width = 22
ws5.column_dimensions["F"].width = 20

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6 – Pirâmide Hierárquica
# ══════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Pirâmide Hierárquica")
ws6.sheet_view.showGridLines = False

ws6.merge_cells("A1:D1")
t = ws6["A1"]
t.value = "PIRÂMIDE HIERÁRQUICA – CARGOS ATIVOS"
t.fill = hdr_fill(DARK_BLUE)
t.font = Font(color=WHITE, bold=True, size=14, name="Calibri")
t.alignment = Alignment(horizontal="center", vertical="center")
ws6.row_dimensions[1].height = 32

headers6 = ["Grupo Hierárquico", "Cargos Ativos", "% do Total", "Ordem"]
for ci, h in enumerate(headers6, 1):
    c = ws6.cell(row=2, column=ci)
    c.value = h
    c.fill = hdr_fill(MID_BLUE)
    c.font = Font(color=WHITE, bold=True, name="Calibri")
    c.alignment = Alignment(horizontal="center")
    c.border = thin_border()
ws6.row_dimensions[2].height = 24

ativos_df = df[df["SITUAÇÃO"] == "ATIVO"]
hier_counts = ativos_df.groupby("GRUPO HIERARQUICO").size()
hier_ordered = [(h, hier_counts.get(h, 0)) for h in hier_order if h in hier_counts.index]
total_ativos = sum(v for _, v in hier_ordered)

for i, (grp, cnt) in enumerate(hier_ordered):
    r = i + 3
    pct = cnt / total_ativos if total_ativos else 0
    for ci, val in enumerate([grp, cnt, f"{pct:.1%}", i + 1], 1):
        c = ws6.cell(row=r, column=ci)
        c.value = val
        c.fill = hdr_fill(ALT_ROW if i % 2 else WHITE)
        c.font = Font(name="Calibri", size=10, bold=(ci == 1))
        c.alignment = Alignment(horizontal="center" if ci > 1 else "left")
        c.border = thin_border()
    ws6.row_dimensions[r].height = 22

# total row
tot_r = len(hier_ordered) + 3
ws6.cell(row=tot_r, column=1).value = "TOTAL"
ws6.cell(row=tot_r, column=2).value = total_ativos
ws6.cell(row=tot_r, column=3).value = "100,0%"
ws6.cell(row=tot_r, column=4).value = ""
apply_header_row(ws6, tot_r, 4, DARK_BLUE)

# Horizontal bar chart
chart6 = BarChart()
chart6.type      = "bar"       # horizontal
chart6.title     = "Pirâmide de Cargos Ativos por Nível Hierárquico"
chart6.x_axis.title = "Quantidade"
chart6.style     = 10
chart6.width     = 20
chart6.height    = 14

data6 = Reference(ws6, min_col=2, max_col=2, min_row=2, max_row=len(hier_ordered) + 2)
cats6 = Reference(ws6, min_col=1, min_row=3, max_row=len(hier_ordered) + 2)
chart6.add_data(data6, titles_from_data=True)
chart6.set_categories(cats6)
chart6.series[0].graphicalProperties.solidFill = MID_BLUE
ws6.add_chart(chart6, "F2")

ws6.column_dimensions["A"].width = 22
ws6.column_dimensions["B"].width = 16
ws6.column_dimensions["C"].width = 14
ws6.column_dimensions["D"].width = 10

# ── save ──────────────────────────────────────────────────────────────────────
OUTPUT = "/home/user/Projects/Dashboard_Cargos_GBS.xlsx"
wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
